import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# 폴더 경로 설정
OUTPUT_DIR = "output"  # 병합할 엑셀 파일들이 저장된 폴더


def merge_excel_files_with_border():
    # 병합할 엑셀 파일들을 가져옵니다.
    excel_files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith(".xlsx")]

    if not excel_files:
        print("병합할 엑셀 파일이 없습니다.")
        return

    # 첫 번째 파일 이름 가져오기
    first_file_name = os.path.splitext(excel_files[0])[0]
    merged_file = f"{first_file_name}_merged_output.xlsx"

    # 첫 번째 파일의 데이터프레임을 읽습니다.
    combined_df = pd.DataFrame()

    for i, excel_file in enumerate(excel_files):
        file_path = os.path.join(OUTPUT_DIR, excel_file)

        # 엑셀 파일 읽기
        df = pd.read_excel(file_path, header=None)

        # 병합 데이터프레임에 파일 이름과 구분을 위한 빈 행 추가
        if not combined_df.empty:
            combined_df = pd.concat([combined_df, pd.DataFrame(
                [[""] * df.shape[1]])], ignore_index=True)

        combined_df = pd.concat([combined_df, df], ignore_index=True)

    # 병합된 데이터프레임을 저장
    output_file_path = os.path.join(OUTPUT_DIR, merged_file)
    combined_df.to_excel(output_file_path, index=False, header=False)

    print(f"병합된 파일을 저장했습니다: {output_file_path}")

    # 테두리를 추가하는 작업
    add_borders_to_excel(output_file_path, combined_df.shape[0])


def add_borders_to_excel(excel_file, total_rows):
    # 얇은 테두리 설정
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 엑셀 파일 불러오기
    wb = load_workbook(excel_file)
    ws = wb.active

    # 각 셀에 테두리 추가 (병합된 모든 행에 대해)
    for row in ws.iter_rows(min_row=1, max_row=total_rows, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # 수정된 엑셀 파일 저장
    wb.save(excel_file)
    print(f"테두리를 추가한 엑셀 파일을 저장했습니다: {excel_file}")


if __name__ == "__main__":
    merge_excel_files_with_border()
